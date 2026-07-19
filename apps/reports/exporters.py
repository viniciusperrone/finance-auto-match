import io
from collections import Counter
from decimal import Decimal

from django.db.models import QuerySet
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors, pagesizes
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from apps.reconciliation.models import ReconciliationResult


Status = ReconciliationResult.Status
STATUS_LABELS = dict(Status.choices)

HEADER_BG_HEX = "1F2937"
STATUS_BG_HEX = {
    Status.CONCILIADO: "D1FAE5",
    Status.DIVERGENCIA: "FEF3C7",
    Status.NAO_ENCONTRADO: "E5E7EB",
    Status.POSSIVEL_DUPLICADO: "FEE2E2",
}


def format_currency_brl(value) -> str:
    if value is None:
        return "—"
    text = f"{Decimal(value):,.2f}"
    return "R$ " + text.replace(",", "X").replace(".", ",").replace("X", ".")

def build_summary(queryset) -> dict:
    counts = Counter(queryset.values_list("status", flat=True))

    return {
        "total": sum(counts.values()),
        "reconciled": counts.get(Status.CONCILIADO, 0),
        "discrepancy": counts.get(Status.DIVERGENCIA, 0),
        "not_found": counts.get(Status.NAO_ENCONTRADO, 0),
        "possible_duplicate": counts.get(Status.POSSIVEL_DUPLICADO, 0)
    }

def export_to_excel(queryset: QuerySet[ReconciliationResult]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Conciliação"

    headers = [
        "Cliente", "Documento", "Vencimento", "Valor esperado", "Status",
        "Score", "Transação encontrada", "Data da transação",
        "Valor da transação", "Diferença de valor", "Diferença de data (dias)",
        "Observações", "Arquivos de origem"
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color=HEADER_BG_HEX, end_color=HEADER_BG_HEX, fill_type="solid")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    row_idx = 2
    for result in queryset.select_related("receivable", "bank_transaction", "receivable__uploaded_file"):
        receivable = result.receivable
        transaction = result.bank_transaction

        ws.append([
            receivable.client_name,
            receivable.document_number,
            receivable.due_date,
            float(receivable.expected_amount),
            STATUS_LABELS.get(result.status, result.status),
            float(result.score),
            transaction.description if transaction else "",
            transaction.transaction_date if transaction else None,
            float(transaction.amount) if transaction else None,
            float(result.amount_difference) if result.amount_difference is not None else None,
            result.date_difference_days,
            result.notes,
            receivable.uploaded_file.original_filename
        ])

        fill_hex = STATUS_BG_HEX.get(result.status)
        if fill_hex:
            fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).font = Font(name="Arial", size=10)
        row_idx += 1

    last_row = row_idx - 1
    for col in (3, 8):
        for row in range(2, row_idx):
            ws.cell(row=row, column=col).number_format = "DD/MM/YYYY"

    for col in (4, 9, 10):
        for row in range(2, row_idx):
            ws.cell(row=row, column=col).number_format = "#,##0.00"

    widths = [24, 14, 12, 14, 22, 8, 14, 14, 14, 12, 50, 26]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    if last_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_row}"

    ws2 = wb.create_sheet("Resumo")
    ws2.append(["Indicador", "Valor"])
    for col in ("A", "B"):
        ws2[f"{col}1"].font = Font(name="Arial", bold=True, color="FFFFFF")
        ws2[f"{col}1"].fill = header_fill

    if last_row >= 2:
        resume_rows = [
            ("Total de registros processados", f"=COUNTA('Conciliação'!A2:A{last_row})"),
            ("Total conciliado", f"=COUNTIF('Conciliação'!E2:E{last_row},\Conciliado\")"),
            ("Total com divergência", f"=COUNTIF('Conciliação'!E2:E{last_row},\Divergência\")"),
            ("Total pendente (não encontro)", f"=COUNTIF('Conciliação'!E2:E{last_row},\Pagamento não encontrado\")"),
            ("Possíveis duplicidades", f"=COUNTIF('Conciliação'!E2:E{last_row},\Possível pagamento duplicado\")")
        ]

    else:
        summary = build_summary(queryset)
        resume_rows = [
            ("Total de registros processados", summary["total"]),
            ("Total conciliado", summary["reconciled"]),
            ("Total com divergência", summary["discrepancy"]),
            ("Total pendente (não encontrado)", summary["not_found"]),
            ("Possíveis duplicidades", summary["possible_duplicate"]),
        ]

    resume_rows.append(("Gerado em", timezone.localtime().strftime("%d/%m/%Y %H:%M")))

    for label, value in resume_rows:
        ws2.append([label, value])

    for row in range(2, len(resume_rows) + 2):
        ws2.cell(row=row, column=1).font = Font(name="Arial", size=10)
        ws2.cell(row=row, column=2).font = Font(name="Arial", size=10)

    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)

    return buffer.getvalue()

def export_to_pdf(queryset) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesizes=landscape(A4),
        leftMargin=1.3 * cm,
        rightMargin=1.3 * cm,
        topMargin=1.3 * cm,
        bottomMargin=1.3 * cm,
        title="Relatório de Conciliação Financeira"
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], fontSize=16, spaceAfter=2)
    subtitle_style = ParagraphStyle("SubtitleCustom", parent=styles["Normal"], fontSize=9, textColor=colors.grey)
    section_style = ParagraphStyle("SectionCustom", parent=styles["Normal"], fontSize=7.5, spaceBefore=14, spaceAfter=6)
    cell_style = ParagraphStyle("CellCustom", parent=styles["Normal"], fontSize=7.5, leading=9)

    story = [
        Paragraph("Relatório de Conciliação Financeira", title_style),
        Paragraph(f"Gerado em {timezone.localtime().strftime('%d/%m/%Y %H:%M')}", subtitle_style),
        Spacer(1, 12)
    ]

    summary = build_summary(queryset)
    resume_data = [
        ["Indicador", "Valor"],
        ["Total de registros processados", str(summary["total"])],
        ["Total conciliação", str(summary["reconciled"])],
        ["Total com divergência", str(summary["discrepancy"])],
        ["Total pendente (não encontrado)", str(summary["not_found"])],
        ["Possíveis duplicidades", str(summary["possible_duplicate"])],
    ]
    resume_table = Table(resume_data, colWidths=[8 * cm, 4 * cm])
    resume_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_BG_HEX}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, 0), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAF8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(resume_table)
    story.append(Paragraph("Detalhamento", section_style))

    table_data = [["Cliente", "Vencimento", "Valor esperado", "Status", "Score", "Transação encontrada", "Observações"]]
    row_statuses = []
    for result in queryset.select_related("receivable", "bank_transaction"):
        receivable = result.receivable
        transaction = result.bank_transaction
        transaction_str = f"{transaction.transaction_date.strftime('%d/%m/%Y')} — {transaction.description}" if transaction else "—"
        table_data.append([
            Paragraph(receivable.client_name, cell_style),
            receivable.due_date.strftime("%d/%m/%Y"),
            format_currency_brl(receivable.expected_amount),
            STATUS_LABELS.get(result.status, result.status),
            f"{result.score:.0f}",
            Paragraph(transaction_str, cell_style),
            Paragraph(result.notes, cell_style),
        ])
        row_statuses.append(result.status)

    if len(table_data) == 1:
        story.append(Paragraph("Nenhum registro encontrado para os filtros aplicados.", styles["Normal"]))
    else:
        detail_table = Table(
            table_data,
            colWidths=[3.3 * cm, 2.1 * cm, 2.6 * cm, 3.0 * cm, 1.3 * cm, 5.5 * cm, 8.2 * cm],
            repeatRows=1,
        )
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_BG_HEX}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("FONTSIZE", (0, 1), (-1, -1), 7.5),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
        for idx, status in enumerate(row_statuses, start=1):
            hex_color = STATUS_BG_HEX.get(status)
            if hex_color:
                style_commands.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor(f"#{hex_color}")))
        detail_table.setStyle(TableStyle(style_commands))
        story.append(detail_table)

    doc.build(story)

    return buffer.getvalue()
